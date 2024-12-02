import openpyxl
from archicad import ACConnection
import datetime

# Verbindung zu Archicad herstellen
conn = ACConnection.connect()
assert conn, "Keine Verbindung zu ARCHICAD möglich. Bitte sicherstellen, dass ARCHICAD läuft."

# Zugriff auf die Module für Befehle und Typen
acc = conn.commands
act = conn.types
acu = conn.utilities

# Alle Stützen-Elemente abrufen
columns = acc.GetElementsByType("Column")

# Built-in Property ID für die Element-ID abrufen
element_id_property_id = acu.GetBuiltInPropertyId('General_ElementID')

# Property-Werte für die abgerufenen Stützen-Elemente erhalten
property_values = acc.GetPropertyValuesOfElements(columns, [element_id_property_id])

# Bounding Boxes der Stützen abrufen
bounding_boxes = acc.Get3DBoundingBoxes(columns)

# Projektinformationen aus Archicad abrufen
project_info = acc.GetProjectInfo()
project_name = project_info.get('ProjectName', 'Unbekanntes Projekt')

# Vorhandene Excel-Datei laden
template_path = 'Stuetzen_Vorlage.xlsx'
wb = openpyxl.load_workbook(template_path)
ws = wb.active

# Datum und Projektname in der Vorlage aktualisieren
current_date = datetime.datetime.now().strftime('%d.%m.%Y')
ws['B6'] = current_date
ws['B7'] = project_name

# Annahme: Die Vorlage hat Kopfzeilen ab Zeile 10 und Daten werden ab Zeile 11 eingefügt
row = 11
for prop, bounding_box in zip(property_values, bounding_boxes):
    if prop.propertyValues[0].propertyValue.value == "Baugespann":
        element_id = prop.propertyValues[0].propertyValue.value
        x_coord = round(bounding_box.boundingBox3D.xMin, 2)
        y_coord = round(bounding_box.boundingBox3D.yMin, 2)
        z_min = round(bounding_box.boundingBox3D.zMin, 2)
        z_max = round(bounding_box.boundingBox3D.zMax, 2)
        height = round(z_max - z_min, 2)

        # Daten in die Excel-Tabelle schreiben
        ws[f'A{row}'] = element_id
        ws[f'B{row}'] = x_coord
        ws[f'C{row}'] = y_coord
        ws[f'D{row}'] = z_min
        ws[f'E{row}'] = height

        row += 1

# Geänderte Excel-Datei speichern
wb.save('Stuetzen_Ausschreibung.xlsx')

# Gesamtergebnis ausgeben
print(f"Excel-Liste 'Stuetzen_Ausschreibung.xlsx' wurde erfolgreich erstellt.")

